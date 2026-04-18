from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from datetime import date, timedelta, datetime
from collections import defaultdict
import csv, io

from sites_mgmt.models import HotspotSite, VoucherTier
from unifi_api import client as unifi


def _get_report_data(request, site_id, date_from_str, date_to_str):
    """
    Fetch voucher-guest sessions from UniFi for the given site(s) and period.
    Uses /stat/guest (same source as dashboard) so figures match exactly.
    Returns (guests_list, site_label).
    """
    date_from_ts = datetime.fromisoformat(date_from_str).timestamp()
    date_to_ts   = (datetime.fromisoformat(date_to_str) + timedelta(days=1)).timestamp()

    if site_id:
        qs = HotspotSite.objects.filter(unifi_site_id=site_id, is_active=True)
        if not request.user.is_superadmin:
            qs = qs.filter(pk__in=request.user.managed_sites.values_list('pk', flat=True))
    elif request.user.is_superadmin:
        qs = HotspotSite.objects.filter(is_active=True)
    else:
        qs = request.user.managed_sites.filter(is_active=True)

    sites_list  = list(qs)
    site_label  = sites_list[0].name if len(sites_list) == 1 else 'Tous les sites'

    tiers = list(VoucherTier.objects.filter(is_active=True).order_by('min_minutes'))

    def tier_for(minutes):
        for t in tiers:
            if t.min_minutes <= minutes <= t.max_minutes:
                return t
        return None

    all_guests = unifi.get_all_guests(sites_list)
    guests = [g for g in all_guests if date_from_ts <= g['sold_ts'] < date_to_ts]

    for g in guests:
        t = tier_for(g['duration_minutes'])
        g['tier_label'] = t.label if t else 'Sans tranche'
        g['price']      = float(t.price_htg) if t else 0

    guests.sort(key=lambda g: g['sold_ts'], reverse=True)
    return guests, site_label


@login_required
def report_index(request):
    today = date.today()
    sites = HotspotSite.objects.filter(is_active=True) if request.user.is_superadmin \
        else request.user.managed_sites.filter(is_active=True)
    return render(request, 'reports/index.html', {
        'sites': sites,
        'today': today,
        'page_title': 'Rapports & Exports',
    })


# ─── CSV ──────────────────────────────────────────────────────────────────────

@login_required
def export_csv(request):
    site_id   = request.GET.get('site', '')
    date_from = request.GET.get('from', (date.today() - timedelta(days=30)).isoformat())
    date_to   = request.GET.get('to',   date.today().isoformat())

    guests, site_label = _get_report_data(request, site_id or None, date_from, date_to)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="bonnet_rapport_{date_from}_{date_to}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Site', 'Forfait', 'Durée (h)', 'Prix (HTG)', 'MAC client', 'Date activation', 'Date expiry'])
    for g in guests:
        writer.writerow([
            g['site_name'],
            g['tier_label'],
            round(g['duration_minutes'] / 60, 1),
            g['price'],
            g.get('mac', '-'),
            g['sold_dt'].strftime('%Y-%m-%d %H:%M') if g['sold_dt'] else '-',
            datetime.fromtimestamp(g['end']).strftime('%Y-%m-%d %H:%M') if g.get('end') else '-',
        ])
    return response


# ─── EXCEL ────────────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    site_id   = request.GET.get('site', '')
    date_from = request.GET.get('from', (date.today() - timedelta(days=30)).isoformat())
    date_to   = request.GET.get('to',   date.today().isoformat())

    guests, site_label = _get_report_data(request, site_id or None, date_from, date_to)

    total_revenue = sum(g['price'] for g in guests)
    now_ts        = datetime.now().timestamp()
    active_count  = sum(1 for g in guests if g.get('end', 0) > now_ts)

    BLUE_DARK = "1E40AF"
    BLUE_LIGHT = "EFF6FF"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor=BLUE_DARK)
    alt   = PatternFill("solid", fgColor=BLUE_LIGHT)

    wb = openpyxl.Workbook()

    # ── Feuille 1 : Résumé ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.merge_cells('A1:E1')
    ws1['A1'] = f"BonNet — Rapport | {site_label}"
    ws1['A1'].font = Font(bold=True, size=14, color=BLUE_DARK)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:E2')
    ws1['A2'] = f"Période : {date_from} → {date_to}"
    ws1['A2'].font = Font(color="6B7280", size=10)
    ws1['A2'].alignment = Alignment(horizontal='center')

    for col, hdr in enumerate(['Indicateur', 'Valeur'], start=1):
        c = ws1.cell(row=4, column=col, value=hdr)
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')

    summary_rows = [
        ('Sessions vendues',                  len(guests)),
        ('Sessions actives (live)',            active_count),
        ('Revenu total (HTG)',                 total_revenue),
        ('Revenu moyen / session (HTG)',       total_revenue / len(guests) if guests else 0),
    ]
    for i, (label, val) in enumerate(summary_rows, start=5):
        ws1.cell(row=i, column=1, value=label)
        cell = ws1.cell(row=i, column=2, value=round(float(val), 2) if isinstance(val, float) else val)
        if 'HTG' in label:
            cell.number_format = '#,##0.00 "HTG"'
        if i % 2 == 0:
            ws1.cell(row=i, column=1).fill = alt
            ws1.cell(row=i, column=2).fill = alt
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 22

    # ── Feuille 2 : Détail sessions ───────────────────────────────────────────
    ws2 = wb.create_sheet("Détail sessions")
    headers    = ['Site', 'Forfait', 'Durée (h)', 'Prix (HTG)', 'MAC client', 'Date activation', 'Date expiry']
    col_widths = [28, 22, 12, 14, 20, 22, 22]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 24
    ws2.freeze_panes = 'A2'

    for ri, g in enumerate(guests, start=2):
        row = [
            g['site_name'],
            g['tier_label'],
            round(g['duration_minutes'] / 60, 1),
            g['price'],
            g.get('mac', '-'),
            g['sold_dt'].strftime('%d/%m/%Y %H:%M') if g['sold_dt'] else '-',
            datetime.fromtimestamp(g['end']).strftime('%d/%m/%Y %H:%M') if g.get('end') else '-',
        ]
        fill = alt if ri % 2 == 0 else None
        for ci, val in enumerate(row, start=1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            if ci == 4:
                cell.number_format = '#,##0.00'

    # ── Feuille 3 : Par forfait ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Par forfait")
    tier_totals = defaultdict(lambda: {'count': 0, 'revenue': 0.0})
    for g in guests:
        tier_totals[g['tier_label']]['count']   += 1
        tier_totals[g['tier_label']]['revenue']  += g['price']

    for col, hdr in enumerate(['Forfait', 'Sessions', 'Revenu (HTG)'], start=1):
        c = ws3.cell(row=1, column=col, value=hdr)
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal='center')

    tier_rows = sorted(tier_totals.items(), key=lambda x: -x[1]['revenue'])
    for i, (label, data) in enumerate(tier_rows, start=2):
        ws3.cell(row=i, column=1, value=label)
        ws3.cell(row=i, column=2, value=data['count'])
        c = ws3.cell(row=i, column=3, value=round(data['revenue'], 2))
        c.number_format = '#,##0.00 "HTG"'
    for col in range(1, 4):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    if tier_rows:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Revenus par forfait"
        chart.y_axis.title = "HTG"
        n = len(tier_rows)
        chart.add_data(Reference(ws3, min_col=3, min_row=1, max_row=1 + n), titles_from_data=True)
        chart.set_categories(Reference(ws3, min_col=1, min_row=2, max_row=1 + n))
        chart.width = 18; chart.height = 12
        ws3.add_chart(chart, "E3")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="bonnet_rapport_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


# ─── PDF ──────────────────────────────────────────────────────────────────────

@login_required
def export_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable

    site_id   = request.GET.get('site', '')
    date_from = request.GET.get('from', (date.today() - timedelta(days=30)).isoformat())
    date_to   = request.GET.get('to',   date.today().isoformat())

    guests, site_label = _get_report_data(request, site_id or None, date_from, date_to)

    total_revenue = sum(g['price'] for g in guests)
    now_ts        = datetime.now().timestamp()
    active_count  = sum(1 for g in guests if g.get('end', 0) > now_ts)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 textColor=colors.HexColor('#1A56DB'), fontSize=16)
    elements = []

    elements.append(Paragraph(f"BonNet — Rapport Vouchers | {site_label}", title_style))
    elements.append(Paragraph(
        f"Période : {date_from} → {date_to}  |  Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal'],
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#3B82F6')))
    elements.append(Spacer(1, 5*mm))

    kpi_data = [
        ['Sessions vendues', 'Sessions actives (live)', 'Revenu total (HTG)'],
        [str(len(guests)), str(active_count), f"{total_revenue:,.2f} HTG"],
    ]
    kpi_table = Table(kpi_data, colWidths=[7*cm, 7*cm, 7*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 13),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#EBF5FF')]),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 6*mm))

    data = [['Site', 'Forfait', 'Durée', 'Prix HTG', 'MAC', 'Date activation']]
    for g in guests[:500]:
        data.append([
            g['site_name'][:25],
            g['tier_label'][:18],
            f"{round(g['duration_minutes']/60,1)}h",
            f"{g['price']:,.0f}",
            g.get('mac', '-'),
            g['sold_dt'].strftime('%d/%m/%Y %H:%M') if g['sold_dt'] else '-',
        ])

    table = Table(data, colWidths=[5.5*cm, 4*cm, 2.2*cm, 2.5*cm, 3.5*cm, 4*cm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A56DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F9FAFB'), colors.white]),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)

    elements.append(Spacer(1, 4*mm))
    total_row = Table(
        [['', '', '', f"TOTAL : {total_revenue:,.2f} HTG", '', '']],
        colWidths=[5.5*cm, 4*cm, 2.2*cm, 2.5*cm, 3.5*cm, 4*cm],
    )
    total_row.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (3, 0), (3, 0), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#DBEAFE')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(total_row)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bonnet_rapport_{date_from}_{date_to}.pdf"'
    return response
