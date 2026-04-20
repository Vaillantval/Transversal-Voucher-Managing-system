"""
Générateurs de rapports : PDF (ReportLab) et Excel (openpyxl).
"""
import io
from datetime import date
from decimal import Decimal

from django.utils import timezone


# ─── PDF ──────────────────────────────────────────────────────────────────────

def generate_pdf_report(queryset, site_name: str, date_from: date, date_to: date) -> bytes:
    """
    Génère un rapport PDF des vouchers.
    Retourne les bytes du fichier PDF.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=20*mm,
        bottomMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=4)
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.gray)
    header_style = ParagraphStyle('H', parent=styles['Normal'], fontSize=9, textColor=colors.white)

    # Calcul totaux
    total_revenue = sum(v.price_htg for v in queryset)
    total_count = queryset.count()
    used_count = queryset.filter(status='used').count()

    elements = []

    # ── Logos ──
    try:
        from sites_mgmt.models import SiteConfig
        from reportlab.platypus import Image as RLImage
        cfg = SiteConfig.get()
        logo_cells = []
        for logo_field in (cfg.logo1, cfg.logo2):
            if logo_field and logo_field.name:
                try:
                    img = RLImage(logo_field.path, height=18*mm, width=50*mm, kind='proportional')
                    logo_cells.append(img)
                except Exception:
                    pass
        if logo_cells:
            logo_table = Table([logo_cells], colWidths=[55*mm] * len(logo_cells))
            logo_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(logo_table)
    except Exception:
        pass

    # ── En-tête ──
    elements.append(Paragraph("BonNet — Rapport de Vouchers", title_style))
    elements.append(Paragraph(
        f"Site : {site_name}  |  Période : {date_from} → {date_to}  |  Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        subtitle_style,
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#3B82F6')))
    elements.append(Spacer(1, 6*mm))

    # ── Résumé ──
    summary_data = [
        ['Total vouchers', 'Vouchers utilisés', 'Revenu total (HTG)'],
        [str(total_count), str(used_count), f"{total_revenue:,.2f} HTG"],
    ]
    summary_table = Table(summary_data, colWidths=[80*mm, 80*mm, 80*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, 1), 13),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#EFF6FF'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8*mm))

    # ── Tableau détaillé ──
    col_headers = ['Code', 'Site', 'Durée', 'Tarif', 'Prix (HTG)', 'Statut', 'Créé le', 'Note']
    rows = [col_headers]
    for v in queryset:
        rows.append([
            v.code,
            v.site.name,
            f"{v.duration_hours}h",
            v.tier.label if v.tier else '—',
            f"{v.price_htg:,.2f}",
            v.get_status_display(),
            v.created_at.strftime('%d/%m/%Y %H:%M'),
            v.note or '—',
        ])

    col_widths = [32*mm, 35*mm, 20*mm, 30*mm, 28*mm, 22*mm, 38*mm, 35*mm]
    detail_table = Table(rows, colWidths=col_widths, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#EFF6FF'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#BFDBFE')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(detail_table)

    # ── Ligne total ──
    elements.append(Spacer(1, 4*mm))
    total_row = Table(
        [['', '', '', 'TOTAL', f"{total_revenue:,.2f} HTG", '', '', '']],
        colWidths=col_widths,
    )
    total_row.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (4, 0), (4, 0), 'RIGHT'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#DBEAFE')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(total_row)

    doc.build(elements)
    return buffer.getvalue()


# ─── EXCEL ────────────────────────────────────────────────────────────────────

def generate_excel_report(queryset, site_name: str, date_from: date, date_to: date) -> bytes:
    """
    Génère un rapport Excel multi-feuilles.
    Feuille 1 : Résumé | Feuille 2 : Détail vouchers | Feuille 3 : Par tranche
    """
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()

    BLUE_DARK = "1E40AF"
    BLUE_LIGHT = "DBEAFE"
    BLUE_MID = "93C5FD"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=BLUE_DARK)
    subheader_fill = PatternFill("solid", fgColor=BLUE_LIGHT)
    thin_border = Border(
        left=Side(style='thin', color=BLUE_MID),
        right=Side(style='thin', color=BLUE_MID),
        top=Side(style='thin', color=BLUE_MID),
        bottom=Side(style='thin', color=BLUE_MID),
    )

    # ── Feuille 1 : Résumé ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Résumé"

    ws1.merge_cells('A1:F1')
    ws1['A1'] = f"BonNet — Rapport Vouchers | {site_name}"
    ws1['A1'].font = Font(bold=True, size=14, color=BLUE_DARK)
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A2:F2')
    ws1['A2'] = f"Période : {date_from} → {date_to}"
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1['A2'].font = Font(color="6B7280", size=10)

    ws1.row_dimensions[4].height = 24
    for col, hdr in enumerate(['Indicateur', 'Valeur'], start=1):
        c = ws1.cell(row=4, column=col, value=hdr)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')

    total_revenue = sum(v.price_htg for v in queryset)
    total_count = queryset.count()
    used_count = queryset.filter(status='used').count()
    pending_count = queryset.filter(status='pending').count()

    summary_rows = [
        ('Total vouchers créés', total_count),
        ('Vouchers utilisés', used_count),
        ('Vouchers en attente', pending_count),
        ('Revenu total (HTG)', float(total_revenue)),
        ('Revenu moyen par voucher (HTG)', float(total_revenue / total_count) if total_count else 0),
    ]
    for i, (label, val) in enumerate(summary_rows, start=5):
        ws1.cell(row=i, column=1, value=label).border = thin_border
        cell = ws1.cell(row=i, column=2, value=val)
        cell.border = thin_border
        if 'HTG' in label:
            cell.number_format = '#,##0.00 "HTG"'
        if i % 2 == 0:
            ws1.cell(row=i, column=1).fill = PatternFill("solid", fgColor="F0F7FF")
            ws1.cell(row=i, column=2).fill = PatternFill("solid", fgColor="F0F7FF")

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 22

    # ── Feuille 2 : Détail ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Détail vouchers")
    headers = ['Code', 'Site', 'Durée (min)', 'Durée (h)', 'Tranche', 'Prix (HTG)', 'Statut', 'Créé par', 'Créé le', 'Note']
    col_widths = [14, 22, 14, 12, 20, 14, 14, 20, 20, 25]

    for col, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws2.cell(row=1, column=col, value=hdr)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = 'A2'

    for row_idx, v in enumerate(queryset, start=2):
        row_data = [
            v.code,
            v.site.name,
            v.duration_minutes,
            v.duration_hours,
            v.tier.label if v.tier else '—',
            float(v.price_htg),
            v.get_status_display(),
            str(v.created_by) if v.created_by else '—',
            v.created_at.strftime('%d/%m/%Y %H:%M'),
            v.note or '',
        ]
        fill = PatternFill("solid", fgColor="EFF6FF") if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if fill:
                cell.fill = fill
            if col_idx == 6:
                cell.number_format = '#,##0.00'

    # ── Feuille 3 : Par tranche ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Par tranche tarifaire")

    from django.db.models import Sum, Count
    tier_data = list(
        queryset.values('tier__label')
        .annotate(count=Count('id'), revenue=Sum('price_htg'))
        .order_by('-count')
    )

    for col, hdr in enumerate(['Tranche', 'Nb vouchers', 'Revenu (HTG)'], start=1):
        c = ws3.cell(row=1, column=col, value=hdr)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    for i, td in enumerate(tier_data, start=2):
        ws3.cell(row=i, column=1, value=td['tier__label'] or 'Non classé')
        ws3.cell(row=i, column=2, value=td['count'])
        c = ws3.cell(row=i, column=3, value=float(td['revenue'] or 0))
        c.number_format = '#,##0.00 "HTG"'

    for col in range(1, 4):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    # Graphique barres
    if len(tier_data) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Revenus par tranche"
        chart.y_axis.title = "HTG"
        chart.x_axis.title = "Tranche"
        data_ref = Reference(ws3, min_col=3, min_row=1, max_row=1 + len(tier_data))
        cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=1 + len(tier_data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 18
        chart.height = 12
        ws3.add_chart(chart, f"E3")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─── CSV ──────────────────────────────────────────────────────────────────────

def generate_csv_report(queryset) -> str:
    """Retourne le contenu CSV sous forme de string."""
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Code', 'Site', 'Durée (min)', 'Tranche', 'Prix HTG',
        'Statut', 'Créé par', 'Créé le', 'Note',
    ])
    for v in queryset:
        writer.writerow([
            v.code,
            v.site.name,
            v.duration_minutes,
            v.tier.label if v.tier else '',
            str(v.price_htg),
            v.get_status_display(),
            str(v.created_by) if v.created_by else '',
            v.created_at.strftime('%Y-%m-%d %H:%M'),
            v.note or '',
        ])
    return output.getvalue()
