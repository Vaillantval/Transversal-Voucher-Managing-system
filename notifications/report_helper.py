"""Génère les rapports Excel et PDF sans requête HTTP (pour les jobs schedulés)."""
import io
from collections import defaultdict
from datetime import datetime, timedelta

from sites_mgmt.utils import TZ_HAITI, find_tier
from sites_mgmt.models import VoucherTier
from unifi_api import client as unifi


def _fetch_guests_per_site(sites: list, date_from: str, date_to: str):
    """Retourne {site: [guests]} filtrés sur la période."""
    date_from_ts = datetime.fromisoformat(date_from).replace(tzinfo=TZ_HAITI).timestamp()
    date_to_ts   = (datetime.fromisoformat(date_to).replace(tzinfo=TZ_HAITI) + timedelta(days=1)).timestamp()

    # Tiers par site (M2M)
    _all_tiers = VoucherTier.objects.filter(is_active=True).prefetch_related('sites')
    tiers_by_site_pk = defaultdict(list)
    for _t in _all_tiers:
        for _s in _t.sites.all():
            tiers_by_site_pk[_s.pk].append(_t)
    unifi_to_pk = {s.unifi_site_id: s.pk for s in sites}

    all_guests = unifi.get_all_guests(sites)

    by_site = defaultdict(list)
    for g in all_guests:
        if date_from_ts <= g['sold_ts'] < date_to_ts:
            _spk = unifi_to_pk.get(g.get('site_unifi_id', ''))
            _st = tiers_by_site_pk.get(_spk, []) if _spk else []
            t = find_tier(_st, g['duration_minutes'])
            g['tier_label'] = t.label if t else 'Sans tranche'
            g['price']      = float(t.price_htg) if t else 0
            by_site[g['site_unifi_id']].append(g)

    for lst in by_site.values():
        lst.sort(key=lambda g: g['sold_ts'], reverse=True)

    return by_site


def generate_excel_bytes(sites: list, date_from: str, date_to: str) -> bytes:
    """
    Un seul fichier Excel :
      - Feuille 1 « Résumé global » : tableau récap (sessions + revenu) par site + grand total
      - Une feuille par site avec le détail complet
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    BLUE_DARK  = "1E40AF"
    BLUE_LIGHT = "EFF6FF"
    hf    = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor=BLUE_DARK)
    alt   = PatternFill("solid", fgColor=BLUE_LIGHT)
    total_fill = PatternFill("solid", fgColor="DBEAFE")

    by_site  = _fetch_guests_per_site(sites, date_from, date_to)
    now_ts   = datetime.now(TZ_HAITI).timestamp()

    wb = openpyxl.Workbook()

    # ── Feuille 1 : Résumé global ─────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Résumé global"

    ws_sum.merge_cells('A1:E1')
    ws_sum['A1'] = "BonNet — Rapport mensuel global"
    ws_sum['A1'].font = Font(bold=True, size=14, color=BLUE_DARK)
    ws_sum['A1'].alignment = Alignment(horizontal='center')

    ws_sum.merge_cells('A2:E2')
    ws_sum['A2'] = f"Période : {date_from}  →  {date_to}"
    ws_sum['A2'].font = Font(color="6B7280", size=10)
    ws_sum['A2'].alignment = Alignment(horizontal='center')

    headers = ['Site', 'Sessions vendues', 'Sessions actives', 'Revenu total (HTG)', 'Revenu moyen (HTG)']
    col_w   = [30, 20, 20, 22, 22]
    for col, (h, w) in enumerate(zip(headers, col_w), start=1):
        c = ws_sum.cell(row=4, column=col, value=h)
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal='center')
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    grand_sessions = grand_active = grand_revenue = 0
    chart_data = []

    for ri, site in enumerate(sites, start=5):
        guests  = by_site.get(site.unifi_site_id, [])
        n       = len(guests)
        active  = sum(1 for g in guests if g.get('end', 0) > now_ts)
        revenue = sum(g['price'] for g in guests)
        avg     = revenue / n if n else 0

        grand_sessions += n
        grand_active   += active
        grand_revenue  += revenue
        chart_data.append((site.name, revenue))

        row_fill = alt if ri % 2 == 0 else None
        for ci, val in enumerate([site.name, n, active, round(revenue, 2), round(avg, 2)], start=1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            if row_fill:
                cell.fill = row_fill
            if ci in (4, 5):
                cell.number_format = '#,##0.00 "HTG"'

    # Ligne total
    total_row = len(sites) + 5
    grand_avg = grand_revenue / grand_sessions if grand_sessions else 0
    for ci, val in enumerate(['TOTAL', grand_sessions, grand_active, round(grand_revenue, 2), round(grand_avg, 2)], start=1):
        cell = ws_sum.cell(row=total_row, column=ci, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        if ci in (4, 5):
            cell.number_format = '#,##0.00 "HTG"'

    # Graphique revenus par site
    if chart_data:
        chart = BarChart()
        chart.type  = "col"
        chart.title = "Revenus par site (HTG)"
        chart.y_axis.title = "HTG"
        n = len(sites)
        chart.add_data(Reference(ws_sum, min_col=4, min_row=4, max_row=4 + n), titles_from_data=True)
        chart.set_categories(Reference(ws_sum, min_col=1, min_row=5, max_row=4 + n))
        chart.width = 20; chart.height = 12
        ws_sum.add_chart(chart, f"A{total_row + 3}")

    # ── Une feuille par site ──────────────────────────────────────────────────
    for site in sites:
        guests = by_site.get(site.unifi_site_id, [])
        # Nettoyer le nom de feuille (Excel : max 31 chars, interdit: / \ * ? [ ] :)
        import re
        sheet_name = re.sub(r'[/\\*?\[\]:]', '-', site.name)[:31]
        ws = wb.create_sheet(sheet_name)

        ws.merge_cells('A1:G1')
        ws['A1'] = f"BonNet — {site.name} | {date_from} → {date_to}"
        ws['A1'].font = Font(bold=True, size=12, color=BLUE_DARK)
        ws['A1'].alignment = Alignment(horizontal='center')

        det_headers = ['Forfait', 'Durée (h)', 'Prix (HTG)', 'MAC client', 'Date activation', 'Date expiry']
        det_widths  = [22, 12, 14, 20, 22, 22]
        for col, (h, w) in enumerate(zip(det_headers, det_widths), start=1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = hf; c.fill = hfill
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = 'A4'

        for ri, g in enumerate(guests, start=4):
            row = [
                g['tier_label'],
                round(g['duration_minutes'] / 60, 1),
                g['price'],
                g.get('mac', '-'),
                g['sold_dt'].strftime('%d/%m/%Y %H:%M') if g['sold_dt'] else '-',
                datetime.fromtimestamp(g['end'], tz=TZ_HAITI).strftime('%d/%m/%Y %H:%M') if g.get('end') else '-',
            ]
            fill = alt if ri % 2 == 0 else None
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
                if ci == 3:
                    cell.number_format = '#,##0.00'

        # Total site
        if guests:
            total_r = len(guests) + 4
            site_total = sum(g['price'] for g in guests)
            ws.cell(row=total_r, column=2, value='TOTAL').font = Font(bold=True)
            cell = ws.cell(row=total_r, column=3, value=round(site_total, 2))
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.number_format = '#,##0.00 "HTG"'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_pdf_bytes(sites: list, date_from: str, date_to: str) -> bytes:
    """
    PDF avec :
      - Tableau récap en haut : revenu + sessions par site + grand total
      - Détail des sessions site par site
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable, PageBreak)

    BLUE       = colors.HexColor('#1A56DB')
    BLUE_DARK  = colors.HexColor('#1E40AF')
    BLUE_LIGHT = colors.HexColor('#DBEAFE')
    ROW_ALT    = colors.HexColor('#F0F7FF')
    GREY_LIGHT = colors.HexColor('#F9FAFB')

    by_site = _fetch_guests_per_site(sites, date_from, date_to)
    now_ts  = datetime.now(TZ_HAITI).timestamp()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=1.5*cm,
    )

    styles    = getSampleStyleSheet()
    title_st  = ParagraphStyle('t', parent=styles['Title'],  textColor=BLUE, fontSize=16)
    h2_st     = ParagraphStyle('h2', parent=styles['Heading2'], textColor=BLUE_DARK, fontSize=12, spaceAfter=4)
    sub_st    = ParagraphStyle('sub', parent=styles['Normal'], textColor=colors.HexColor('#6B7280'), fontSize=9)

    elements = []

    # ── Logos ────────────────────────────────────────────────────────────────
    try:
        from sites_mgmt.models import SiteConfig
        from reportlab.platypus import Image as RLImage
        cfg = SiteConfig.get()
        import os
        logo_cells = []
        for logo_field in (cfg.logo1, cfg.logo2):
            if logo_field and logo_field.name:
                try:
                    path = logo_field.path
                    if os.path.isfile(path) and not path.lower().endswith('.svg'):
                        img = RLImage(path, height=18*mm, width=50*mm, kind='proportional')
                        logo_cells.append(img)
                except Exception:
                    pass
        if logo_cells:
            logo_table = Table([logo_cells], colWidths=[55*mm] * len(logo_cells))
            logo_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(logo_table)
    except Exception:
        pass

    # ── En-tête ───────────────────────────────────────────────────────────────
    elements.append(Paragraph("BonNet — Rapport mensuel", title_st))
    elements.append(Paragraph(
        f"Période : {date_from}  →  {date_to}  |  Généré le {datetime.now(TZ_HAITI).strftime('%d/%m/%Y %H:%M')}",
        sub_st,
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=BLUE))
    elements.append(Spacer(1, 6*mm))

    # ── Tableau récap global ──────────────────────────────────────────────────
    elements.append(Paragraph("Récapitulatif par site", h2_st))

    recap_data = [['Site', 'Sessions vendues', 'Sessions actives', 'Revenu total (HTG)']]
    grand_sessions = grand_active = grand_revenue = 0

    for site in sites:
        guests  = by_site.get(site.unifi_site_id, [])
        n       = len(guests)
        active  = sum(1 for g in guests if g.get('end', 0) > now_ts)
        revenue = sum(g['price'] for g in guests)
        grand_sessions += n
        grand_active   += active
        grand_revenue  += revenue
        recap_data.append([site.name, str(n), str(active), f"{revenue:,.2f} HTG"])

    recap_data.append(['TOTAL', str(grand_sessions), str(grand_active), f"{grand_revenue:,.2f} HTG"])

    n_sites    = len(sites)
    recap_table = Table(recap_data, colWidths=[7*cm, 4.5*cm, 4.5*cm, 5*cm], repeatRows=1)
    recap_style = [
        ('BACKGROUND',  (0, 0),            (-1, 0),             BLUE_DARK),
        ('TEXTCOLOR',   (0, 0),            (-1, 0),             colors.white),
        ('FONTNAME',    (0, 0),            (-1, 0),             'Helvetica-Bold'),
        ('ALIGN',       (0, 0),            (-1, -1),            'CENTER'),
        ('ALIGN',       (0, 1),            (0, -1),             'LEFT'),
        ('FONTNAME',    (0, n_sites + 1),  (-1, n_sites + 1),   'Helvetica-Bold'),
        ('BACKGROUND',  (0, n_sites + 1),  (-1, n_sites + 1),   BLUE_LIGHT),
        ('BOX',         (0, 0),            (-1, -1),            0.5, colors.grey),
        ('GRID',        (0, 0),            (-1, -1),            0.25, colors.lightgrey),
        ('TOPPADDING',  (0, 0),            (-1, -1),            5),
        ('BOTTOMPADDING', (0, 0),          (-1, -1),            5),
    ]
    for i in range(1, n_sites + 1):
        if i % 2 == 0:
            recap_style.append(('BACKGROUND', (0, i), (-1, i), ROW_ALT))

    recap_table.setStyle(TableStyle(recap_style))
    elements.append(recap_table)
    elements.append(Spacer(1, 10*mm))

    # ── Détail par site ───────────────────────────────────────────────────────
    det_headers = ['Forfait', 'Durée', 'Prix HTG', 'MAC', 'Date activation', 'Date expiry']
    det_widths  = [4.5*cm, 2.2*cm, 2.8*cm, 3.8*cm, 4.2*cm, 4.2*cm]

    for site in sites:
        guests = by_site.get(site.unifi_site_id, [])
        site_revenue = sum(g['price'] for g in guests)

        elements.append(Paragraph(f"{site.name}", h2_st))
        elements.append(Paragraph(
            f"{len(guests)} session(s)  |  Revenu : {site_revenue:,.2f} HTG", sub_st
        ))
        elements.append(Spacer(1, 3*mm))

        if not guests:
            elements.append(Paragraph("Aucune session sur cette période.", styles['Normal']))
            elements.append(Spacer(1, 6*mm))
            continue

        data = [det_headers]
        for g in guests[:300]:
            data.append([
                g['tier_label'][:20],
                f"{round(g['duration_minutes']/60, 1)}h",
                f"{g['price']:,.0f}",
                g.get('mac', '-'),
                g['sold_dt'].strftime('%d/%m/%Y %H:%M') if g['sold_dt'] else '-',
                datetime.fromtimestamp(g['end'], tz=TZ_HAITI).strftime('%d/%m/%Y %H:%M') if g.get('end') else '-',
            ])

        # Ligne total site
        data.append(['TOTAL', '', f"{site_revenue:,.2f} HTG", '', '', ''])

        n_rows = len(data)
        tbl = Table(data, colWidths=det_widths, repeatRows=1)
        tbl_style = [
            ('BACKGROUND',    (0, 0),          (-1, 0),          BLUE),
            ('TEXTCOLOR',     (0, 0),          (-1, 0),          colors.white),
            ('FONTNAME',      (0, 0),          (-1, 0),          'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0),          (-1, -1),         8),
            ('ALIGN',         (0, 0),          (-1, -1),         'CENTER'),
            ('FONTNAME',      (0, n_rows - 1), (-1, n_rows - 1), 'Helvetica-Bold'),
            ('BACKGROUND',    (0, n_rows - 1), (-1, n_rows - 1), BLUE_LIGHT),
            ('BOX',           (0, 0),          (-1, -1),         0.5, colors.grey),
            ('GRID',          (0, 0),          (-1, -1),         0.25, colors.lightgrey),
            ('TOPPADDING',    (0, 0),          (-1, -1),         3),
            ('BOTTOMPADDING', (0, 0),          (-1, -1),         3),
        ]
        for i in range(1, n_rows - 1):
            if i % 2 == 0:
                tbl_style.append(('BACKGROUND', (0, i), (-1, i), GREY_LIGHT))

        tbl.setStyle(TableStyle(tbl_style))
        elements.append(tbl)
        elements.append(Spacer(1, 8*mm))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_store_weekly_pdf_bytes(date_from: str, date_to: str) -> bytes:
    """PDF du rapport hebdomadaire des ventes store (Orders)."""
    import io
    from datetime import datetime
    from collections import defaultdict
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from store.models import Order, OrderItem
    from sites_mgmt.utils import TZ_HAITI

    BLUE      = colors.HexColor('#1A56DB')
    BLUE_DARK = colors.HexColor('#1E40AF')
    BLUE_LIGHT= colors.HexColor('#DBEAFE')
    ROW_ALT   = colors.HexColor('#F0F7FF')
    GREEN     = colors.HexColor('#065F46')

    dt_from = datetime.fromisoformat(date_from).replace(tzinfo=TZ_HAITI)
    dt_to   = (datetime.fromisoformat(date_to).replace(tzinfo=TZ_HAITI)
               .replace(hour=23, minute=59, second=59))

    orders = list(
        Order.objects.filter(
            created_at__gte=dt_from,
            created_at__lte=dt_to,
            status=Order.STATUS_DELIVERED,
        ).select_related('customer').prefetch_related('items__tier', 'items__site')
        .order_by('-created_at')
    )

    # Agrégats
    total_revenue = sum(o.total_htg for o in orders)
    by_site  = defaultdict(lambda: {'count': 0, 'revenue': 0.0})
    by_tier  = defaultdict(lambda: {'count': 0, 'revenue': 0.0})
    for o in orders:
        for item in o.items.all():
            site_name = item.site.name if item.site else '—'
            tier_label = item.tier.label if item.tier else '—'
            qty = item.quantity
            rev = float(item.unit_price) * qty
            by_site[site_name]['count']   += qty
            by_site[site_name]['revenue'] += rev
            by_tier[tier_label]['count']  += qty
            by_tier[tier_label]['revenue']+= rev

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=1.5*cm,
    )
    styles   = getSampleStyleSheet()
    title_st = ParagraphStyle('t',  parent=styles['Title'],   textColor=BLUE, fontSize=16)
    h2_st    = ParagraphStyle('h2', parent=styles['Heading2'],textColor=BLUE_DARK, fontSize=12, spaceAfter=4)
    sub_st   = ParagraphStyle('sub',parent=styles['Normal'],  textColor=colors.HexColor('#6B7280'), fontSize=9)

    hf    = ('FONTNAME',    (0,0),(-1,0), 'Helvetica-Bold')
    hbg   = ('BACKGROUND',  (0,0),(-1,0), BLUE_DARK)
    hfg   = ('TEXTCOLOR',   (0,0),(-1,0), colors.white)
    base_style = [hf, hbg, hfg,
                  ('ALIGN',(0,0),(-1,-1),'CENTER'),
                  ('FONTSIZE',(0,0),(-1,-1),8),
                  ('BOX',(0,0),(-1,-1),0.5,colors.grey),
                  ('GRID',(0,0),(-1,-1),0.25,colors.lightgrey),
                  ('TOPPADDING',(0,0),(-1,-1),4),
                  ('BOTTOMPADDING',(0,0),(-1,-1),4)]

    elements = []
    elements.append(Paragraph("BonNet — Rapport hebdomadaire des ventes", title_st))
    elements.append(Paragraph(
        f"Période : {date_from}  →  {date_to}  |  Généré le {datetime.now(TZ_HAITI).strftime('%d/%m/%Y %H:%M')}",
        sub_st,
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=BLUE))
    elements.append(Spacer(1, 6*mm))

    # ── KPIs ─────────────────────────────────────────────────────────────────
    kpi_data = [
        ['Commandes livrées', 'Forfaits vendus', 'Revenu total (HTG)'],
        [str(len(orders)),
         str(sum(v['count'] for v in by_tier.values())),
         f"{float(total_revenue):,.2f} HTG"],
    ]
    kpi_table = Table(kpi_data, colWidths=[6*cm, 6*cm, 7*cm])
    kpi_style = base_style + [
        ('FONTNAME',(0,1),(-1,1),'Helvetica-Bold'),
        ('FONTSIZE',(0,1),(-1,1),13),
        ('TEXTCOLOR',(0,1),(-1,1),BLUE_DARK),
        ('BACKGROUND',(0,1),(-1,1),BLUE_LIGHT),
    ]
    kpi_table.setStyle(TableStyle(kpi_style))
    elements.append(kpi_table)
    elements.append(Spacer(1, 8*mm))

    # ── Recap par site ────────────────────────────────────────────────────────
    elements.append(Paragraph("Ventes par site", h2_st))
    site_data = [['Site', 'Forfaits vendus', 'Revenu (HTG)']]
    for sname, vals in sorted(by_site.items()):
        site_data.append([sname, str(vals['count']), f"{vals['revenue']:,.2f} HTG"])
    site_data.append(['TOTAL',
                      str(sum(v['count'] for v in by_site.values())),
                      f"{sum(v['revenue'] for v in by_site.values()):,.2f} HTG"])
    n_sites = len(by_site)
    site_tbl = Table(site_data, colWidths=[8*cm, 5*cm, 6*cm])
    site_style = base_style + [
        ('ALIGN',(0,1),(0,-1),'LEFT'),
        ('FONTNAME',(0,n_sites+1),(-1,n_sites+1),'Helvetica-Bold'),
        ('BACKGROUND',(0,n_sites+1),(-1,n_sites+1),BLUE_LIGHT),
    ]
    for i in range(1, n_sites+1):
        if i % 2 == 0:
            site_style.append(('BACKGROUND',(0,i),(-1,i),ROW_ALT))
    site_tbl.setStyle(TableStyle(site_style))
    elements.append(site_tbl)
    elements.append(Spacer(1, 8*mm))

    # ── Recap par forfait ─────────────────────────────────────────────────────
    elements.append(Paragraph("Ventes par forfait", h2_st))
    tier_data = [['Forfait', 'Qté vendue', 'Revenu (HTG)']]
    for tlabel, vals in sorted(by_tier.items()):
        tier_data.append([tlabel, str(vals['count']), f"{vals['revenue']:,.2f} HTG"])
    n_tiers = len(by_tier)
    tier_tbl = Table(tier_data, colWidths=[8*cm, 5*cm, 6*cm])
    tier_style = base_style + [
        ('ALIGN',(0,1),(0,-1),'LEFT'),
    ]
    for i in range(1, n_tiers+1):
        if i % 2 == 0:
            tier_style.append(('BACKGROUND',(0,i),(-1,i),ROW_ALT))
    tier_tbl.setStyle(TableStyle(tier_style))
    elements.append(tier_tbl)
    elements.append(Spacer(1, 8*mm))

    # ── Détail des commandes ──────────────────────────────────────────────────
    elements.append(Paragraph("Détail des commandes", h2_st))
    if orders:
        det_data = [['Référence', 'Date', 'Client', 'Téléphone', 'Site(s)', 'Montant HTG']]
        for o in orders:
            sites_str = ', '.join(
                {item.site.name for item in o.items.all() if item.site}
            ) or '—'
            det_data.append([
                o.reference,
                o.created_at.astimezone(TZ_HAITI).strftime('%d/%m %H:%M'),
                (o.customer.full_name[:20] if o.customer else '—'),
                (o.customer.phone if o.customer else '—'),
                sites_str[:25],
                f"{float(o.total_htg):,.2f}",
            ])
        det_tbl = Table(det_data, colWidths=[4.5*cm, 2.8*cm, 4*cm, 3.5*cm, 5*cm, 3.5*cm], repeatRows=1)
        det_style = base_style[:]
        n_orders = len(orders)
        for i in range(1, n_orders+1):
            if i % 2 == 0:
                det_style.append(('BACKGROUND',(0,i),(-1,i),ROW_ALT))
        det_tbl.setStyle(TableStyle(det_style))
        elements.append(det_tbl)
    else:
        elements.append(Paragraph("Aucune commande livrée sur cette période.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
